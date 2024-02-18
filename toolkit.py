import datetime
from pathlib import Path
from typing import Any
import warnings
from warnings import warn

from hdbscan import HDBSCAN
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
import plotly.graph_objects as go

# TODO: pandas is updating to 3.0 soon, and importing it currently generates an annoying deprecation
#  warning. Remove this context manager when 3.0 is released and has been tested in April 2024.
with warnings.catch_warnings():
    warnings.simplefilter('ignore')
    import pandas as pd
import seaborn as sns
from sklearn.preprocessing import LabelEncoder
from umap import UMAP


def main():
    """
    Imports six classical databases maintained in the seaborn-data repository and generates summary
    workbooks for them in './testing_output', creating the directory if it does not already exist.
    :return: None
    """

    root_dirpath = Path('./testing_output/')
    root_dirpath.mkdir(exist_ok=True)

    title_dict = {
        'x': 0.9, 'y': 0.5, 'xanchor': 'top', 'yanchor': 'top',
        'font': {
            'size': 24,
            'color': 'Black',
            'family': 'Arial, sans-serif',
            'weight': 'bold'
        }
    }

    # URL of the Iris dataset hosted on seaborn's GitHub repository
    dataset_url = 'https://raw.githubusercontent.com/mwaskom/seaborn-data/master/iris.csv'
    summary_dirpath = root_dirpath.joinpath('iris/')
    summary_dirpath.mkdir(exist_ok=True)
    fig_filepath = summary_dirpath.joinpath('umap_decomposition.html')
    df = pd.read_csv(dataset_url)
    build_dataset_summary_wb(df, summary_dirpath.joinpath('summary.xlsx'), include_raw_data=True)
    fig_stacked_umap_decomposition_2d(
        df, columns=['sepal_length', 'sepal_width', 'petal_length', 'petal_width'],
        figure_kwargs={'layout': {'title': title_dict.update({'title': 'Iris Dataset'})}},
        fig_filepath=fig_filepath, random_seed=42
    )

    # The Titanic dataset
    dataset_url = 'https://raw.githubusercontent.com/mwaskom/seaborn-data/master/titanic.csv'
    summary_dirpath = root_dirpath.joinpath('titanic/')
    summary_dirpath.mkdir(exist_ok=True)
    fig_filepath = summary_dirpath.joinpath('umap_decomposition.html')
    df = pd.read_csv(dataset_url)
    build_dataset_summary_wb(df, summary_dirpath.joinpath('summary.xlsx'), include_raw_data=True)
    fig_stacked_umap_decomposition_2d(
        df, columns=['pclass', 'sex', 'age', 'sibsp', 'parch', 'fare'],
        figure_kwargs={'layout': {'title': title_dict.update({'title': 'Titanic Dataset'})}},
        fig_filepath=fig_filepath, random_seed=42
    )

    # The Penguins dataset
    dataset_url = 'https://raw.githubusercontent.com/mwaskom/seaborn-data/master/penguins.csv'
    summary_dirpath = root_dirpath.joinpath('penguins/')
    summary_dirpath.mkdir(exist_ok=True)
    fig_filepath = summary_dirpath.joinpath('umap_decomposition.html')
    df = pd.read_csv(dataset_url)
    build_dataset_summary_wb(df, summary_dirpath.joinpath('summary.xlsx'), include_raw_data=True)
    fig_stacked_umap_decomposition_2d(
        df, columns=['bill_length_mm', 'bill_depth_mm', 'flipper_length_mm', 'body_mass_g'],
        figure_kwargs={'layout': {'title': title_dict.update({'title': 'Penguins Dataset'})}},
        fig_filepath=fig_filepath, random_seed=42
    )

    # The Exercise dataset
    dataset_url = 'https://raw.githubusercontent.com/mwaskom/seaborn-data/master/exercise.csv'
    summary_dirpath = root_dirpath.joinpath('exercise/')
    summary_dirpath.mkdir(exist_ok=True)
    fig_filepath = summary_dirpath.joinpath('umap_decomposition.html')
    df = pd.read_csv(dataset_url)
    build_dataset_summary_wb(df, summary_dirpath.joinpath('summary.xlsx'), include_raw_data=True)
    fig_stacked_umap_decomposition_2d(
        df, columns=['diet', 'pulse', 'time', 'kind'],
        figure_kwargs={'layout': {'title': title_dict.update({'title': 'Exercise Dataset'})}},
        fig_filepath=fig_filepath, random_seed=42
    )

    # The Flights dataset
    dataset_url = 'https://raw.githubusercontent.com/mwaskom/seaborn-data/master/flights.csv'
    summary_dirpath = root_dirpath.joinpath('flights/')
    summary_dirpath.mkdir(exist_ok=True)
    fig_filepath = summary_dirpath.joinpath('umap_decomposition.html')
    df = pd.read_csv(dataset_url)
    build_dataset_summary_wb(df, summary_dirpath.joinpath('summary.xlsx'), include_raw_data=True)
    fig_stacked_umap_decomposition_2d(
        df, columns=['year', 'month', 'passengers'],
        figure_kwargs={'layout': {'title': title_dict.update({'title': 'Flights Dataset'})}},
        fig_filepath=fig_filepath, random_seed=42
    )


def create_dataframe_from_file(
        filepath: str
) -> pd.DataFrame | None:
    """
    Facilitates passing common data file types from command line.
    :param filepath: Path to file.
    :return: A DataFrame
    """

    # Get file extension
    file_extension = filepath.split('.')[-1].lower()

    # Attempt to read the file based on the extension
    try:
        if file_extension == 'csv':
            return pd.read_csv(filepath)
        elif file_extension in ['xls', 'xlsx']:
            return pd.read_excel(filepath)
        elif file_extension == 'json':
            return pd.read_json(filepath)
        elif file_extension == 'parquet':
            return pd.read_parquet(filepath)
        else:
            print(f"Unsupported file type: .{file_extension}")
            return None
    except Exception as e:
        print(f"Error reading the file: {e}")
        return None


def build_dataset_summary_wb(
        df: pd.DataFrame,
        excel_filepath: Path | str = None,
        overwrite_existing_file: bool = True,
        max_field_count_values: int = 10_000,
        check_fx_corr: bool = True,
        include_raw_data: bool = False
) -> None:
    """
    Generates detailed summary data for the dataset df and stores it in an Excel file.
    :param df: The dataframe to analyze.
    :param excel_filepath: A valid filepath. Missing directories will not be created automatically.
    Existing files will be overwritten by default. Filenames should end in '.xlsx'.
    :param overwrite_existing_file: If true, will overwrite an existing file. If False, and a file already
    exists at the specified destination, an exception is raised.
    :param max_field_count_values: The maximum number of values to include on the "Field Value Counts"
    worksheet. The most frequent max_field_count values will be displayed in a table along with counts.
    Defaults to 10,000. For very large datasets, especially where there is a primary key, this is necessary
    to not exceed the row limit in Excel.
    :param check_fx_corr: Whether to check for functional correlation between columns. Functionally
    correlated columns have 1:1 mappings between values. This can be memory intensive.
    :param include_raw_data: Whether to include the full data set in the summary. False by default.
    :return: None
    """

    # Validate excel_filepath. If no filepath is specified, default to a generic file name in pwd.
    if excel_filepath is None:
        excel_filepath = Path('./dataset_summary.xlsx')
    elif isinstance(excel_filepath, str):
        excel_filepath = Path(excel_filepath)
    elif not isinstance(excel_filepath, Path):
        raise ValueError("Invalid excel_filepath value.")

    # If file already exists and overwrite_existing_file then raise exception.
    if not overwrite_existing_file and excel_filepath.exists():
        raise FileExistsError("A data file with the specified path already exists. Set "
                              "'overwrite_existing_file' to true to overwrite.")

    # Create a workbook to populate, and get first worksheet.
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data Set Summary'

    # Populate first worksheet with summary data for dataset as a whole.
    ws.cell(1, 1, 'Date Generated:')
    cell = ws.cell(1, 2, pd.Timestamp.now())
    cell.number_format = 'YYYY-MM-DD HH:MM:SS'
    ws.cell(2, 1, 'Number of Columns:')
    cell = ws.cell(2, 2, df.shape[1])
    cell.number_format = get_cell_number_format(df.shape[1])
    ws.cell(3, 1, 'Number of Rows: ')
    cell = ws.cell(3, 2, df.shape[0])
    cell.number_format = get_cell_number_format(df.shape[0])

    ws.cell(4, 1, 'Number of Values Missing:')
    value = df.isna().sum().sum()
    cell = ws.cell(4, 2, value)
    cell.number_format = get_cell_number_format(value)
    ws.cell(5, 1, 'Fraction of Values Missing:')
    value = round(df.isna().sum().sum() / (df.shape[0] * df.shape[1]), 4)
    cell = ws.cell(5, 2, value)
    cell.number_format = get_cell_number_format(value)

    ws.cell(6, 1, 'Number of Values Equal to 0:')
    value = (df == 0).sum().sum()
    cell = ws.cell(6, 2, value)
    cell.number_format = get_cell_number_format(value)
    ws.cell(7, 1, 'Fraction of Values Equal to 0:')
    value = round((df == 0).sum().sum() / (df.shape[0] * df.shape[1]), 4)
    cell = ws.cell(7, 2, value)
    cell.number_format = get_cell_number_format(value)

    ws.cell(8, 1, 'Size in Memory as DataFrame (MB):')
    value = round(df.memory_usage(deep=True).sum() / (1024 ** 2), 4)
    cell = ws.cell(8, 2, value)
    cell.number_format = get_cell_number_format(value)

    ws.column_dimensions['A'].width = 35
    bold_font = Font(bold=True)
    for cell in ws['A']:
        cell.font = bold_font
    ws.column_dimensions['B'].width = 25

    # Create worksheet containing summary statistics for each field.
    ws = wb.create_sheet('Field Summaries')
    build_column_summary_ws(df, ws, check_fx_corr)

    # Create worksheet containing correlation values for fields.
    ws = wb.create_sheet('Correlation Matrix')
    build_corr_matrix_ws(df, ws)

    # If there are columns with an object dtype, generate a worksheet containing a breakdown of the
    # subtypes of data and their distribution within the column.
    if 'object' in df.dtypes.tolist():
        ws = wb.create_sheet('Object Dtype Dist')
        build_object_dtype_dist_ws(df, ws)

    # Create a worksheet with tables of value counts for each field. Only include max_field_count_values
    # most frequently occurring values.
    ws = wb.create_sheet('Value Counts')
    build_field_value_count_ws(df, ws, max_field_count_values)

    # Repeat, but with frequencies instead of counts
    ws = wb.create_sheet('Value Frequencies')
    build_field_value_count_ws(df, ws, max_field_count_values, normalize=True)

    if include_raw_data:
        ws = wb.create_sheet('Raw Data')

        # Write the header row
        bold_font = Font(bold=True)
        for column, column_title in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=column, value=column_title)
            cell.font = bold_font
            ws.column_dimensions[get_column_letter(column)].width = 15

        # Write the DataFrame data
        for row, data in enumerate(df.values, start=2):
            for column, value in enumerate(data, start=1):
                ws.cell(row=row, column=column, value=value)

        # Turn it into a table
        table_range = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
        display_name = "areallylongstringthatwouldneverbeusedasacolumnnametoavoidcollisions"
        table = Table(displayName=display_name, ref=table_range)
        ws.add_table(table)

    wb.save(excel_filepath)


def build_column_summary_ws(
        df: pd.DataFrame,
        ws: Worksheet,
        check_fx_cor: bool = True
) -> None:
    """
    Populates the passed worksheet with summary statistics for each field of data in the dataframe.
    :param df: The dataframe to analyze.
    :param ws: The worksheet to populate.
    :param check_fx_cor: Whether to check for functional correlation between columns. Functionally
    correlated columns have 1:1 mappings between values. This can be memory intensive.
    :return: None
    """
    # Get basic summary data applicable to all data types.
    results = get_column_summary(df, check_fx_cor)
    # Get summary data only applicable to numeric data types.
    numerical_results = get_numeric_column_summary(df)
    # Get data on ordering of data in dataframe.
    ordering_results = get_column_ordering(df)

    # Populate column labels
    bold_font = Font(bold=True)
    for i, column in enumerate(df.columns):
        column_number = i + 2
        cell = ws.cell(1, column_number, column)
        cell.font = bold_font
        ws.column_dimensions[get_column_letter(column_number)].width = 20

    # Populate row labels and freeze first column
    ws.column_dimensions['A'].width = 30
    ws.freeze_panes = 'B1'
    row_number = 2
    for row_label in results.keys():
        cell = ws.cell(row_number, 1, row_label)
        cell.font = bold_font
        row_number += 1
    row_number += 1
    for row_label in numerical_results.keys():
        cell = ws.cell(row_number, 1, row_label)
        cell.font = bold_font
        row_number += 1
    row_number += 1
    for row_label in ordering_results.keys():
        cell = ws.cell(row_number, 1, row_label)
        cell.font = bold_font
        row_number += 1

    # Iterate through columns and populate summary statistics for column into ws.
    for i, column in enumerate(df.columns):
        column_number = i + 2
        row_number = 2

        for value in results.values():
            value = value[column]
            if np.issubdtype(type(value), np.number) or isinstance(value, float):
                value = round(value, 4)
            cell = ws.cell(row_number, column_number, value)
            cell.number_format = get_cell_number_format(value)
            row_number += 1

        # Skip columns that do not contain numerical data.
        if column in numerical_results[list(numerical_results.keys())[0]]:
            row_number += 1
            for value in numerical_results.values():
                value = value[column]
                if np.issubdtype(type(value), np.number) or isinstance(value, float):
                    value = round(value, 4)
                cell = ws.cell(row_number, column_number, value)
                cell.number_format = get_cell_number_format(value)
                row_number += 1
        else:
            row_number += 1 + len(numerical_results.keys())

        row_number += 1
        for value in ordering_results.values():
            value = value[column]
            if np.issubdtype(type(value), np.number) or isinstance(value, float):
                value = round(value, 4)
            cell = ws.cell(row_number, column_number, value)
            cell.number_format = get_cell_number_format(value)
            row_number += 1


def build_corr_matrix_ws(
        df: pd.DataFrame,
        ws: Worksheet
) -> None:
    """
    Populates a worksheet with a correlation matrix for the numeric columns of the dataframe.
    Some correlations may not be calculated if all fields are missing data in a column, or if all fields
    contain the same numeric value, in which case the standard deviation is zero. In these cases the
    relevant field in the correlation matrix will be left blank for readability.
    :param df: The dataframe to analyze.
    :param ws: The worksheet to populate.
    :return: None
    """
    # Select only numeric columns.
    df = df.select_dtypes(include=[np.number])
    # Get the correlation matrix.
    cm = df.corr()

    # Populate row and column labels.
    bold_font = Font(bold=True)
    row_number = 1
    column_number = 2
    for column in df.columns:
        cell = ws.cell(row_number, column_number, column)
        cell.font = bold_font
        ws.column_dimensions[get_column_letter(column_number)].width = 10
        column_number += 1
    row_number = 2
    column_number = 1
    ws.column_dimensions['A'].width = 20
    ws.freeze_panes = 'B1'
    for column in df.columns:
        cell = ws.cell(row_number, column_number, column)
        cell.font = bold_font
        row_number += 1

    # Populate data
    for i, column in enumerate(cm.columns):
        for j, row in enumerate(cm.columns):
            value = cm.loc[row, column]
            # Cells with NaN values will be left blank.
            if not np.isnan(value):
                value = round(value, 4)
                cell = ws.cell(j + 2, i + 2, value)
                cell.number_format = get_cell_number_format(value)


def build_object_dtype_dist_ws(
        df: pd.DataFrame,
        ws: Worksheet
) -> None:
    """
    Populates a worksheet with the breakdown of dtypes within columns that are of dtype object. Object is the
    default data type where mixed data is present, and this will show the distribution of that mixed data.
    If there are no columns with object dtype, will not write anything to the worksheet.
    :param df: The dataframe to analyze.
    :param ws: The worksheet to populate.
    :return: None
    """

    # Get the columns whose datatype is object.
    object_columns = [
        column for column in df.columns if df[column].dtype == 'object'
    ]

    # If no columns of dtype object, generate a warning and do nothing.
    if not object_columns:
        warn("No object dtype columns identified in dataframe.")
    else:
        # Get dtype counts.
        dtype_counts = {
            column: get_dtype_counts(df[column]).to_dict()
            for column in object_columns
        }

        # Populate column labels.
        bold_font = Font(bold=True)
        row_number = 1
        column_number = 2
        for column in object_columns:
            cell = ws.cell(row_number, column_number, column)
            cell.font = bold_font
            ws.column_dimensions[get_column_letter(column_number)].width = 15
            column_number += 1
        row_number = 2
        column_number = 1

        # Populate row labels.
        ws.column_dimensions['A'].width = 20
        ws.freeze_panes = 'B1'
        dtypes = set()
        for sub_dict in dtype_counts.values():
            dtypes.update(sub_dict.keys())
        dtypes = sorted(list(dtypes))
        for dtype in dtypes:
            cell = ws.cell(row_number, column_number, dtype)
            cell.font = bold_font
            row_number += 1

        # Populate value counts.
        for column, counts in dtype_counts.items():
            for dtype, count in counts.items():
                cell = ws.cell(dtypes.index(dtype) + 2, object_columns.index(column) + 2, count)
                cell.number_format = get_cell_number_format(count)


def build_field_value_count_ws(
        df: pd.DataFrame,
        ws: Worksheet,
        max_values: int = 10_000,
        normalize: bool = False,
        table_name_id: str = ''
) -> None:
    """
    Populates a worksheet with tables of value counts for each field in the dataframe. Only includes the
    most frequently occurring max_values items. This limit is useful for very large datasets, particularly
    those with primary keys, which might otherwise overflow the maximum number of rows in an Excel
    spreadsheet.
    :param df: The dataframe to analyze.
    :param ws: The worksheet to populate.
    :param max_values: The maximum number of values to include. Will sort by frequency in descending order
    before trimming excess values.
    :param normalize: If true, will populate frequencies instead of counts.
    :param table_name_id: Table names must be unique across all worksheets in a workbook. If necessary,
    a string can be passed to this method that will be prepended to the table names it generates to ensure
    uniqueness. This is not necessary if column names are not duplicated, and should generally not need to
    be used.
    :return: None
    """

    ws.freeze_panes = 'A3'
    bold_font = Font(bold=True)
    center_aligned_text = Alignment(horizontal='center')

    column_number = 1
    for column in df.columns:
        value_counts = df[column].value_counts(normalize=normalize, dropna=False)
        value_counts = value_counts.iloc[:max_values]
        cell = ws.cell(1, column_number, column)
        cell.font = bold_font
        cell.alignment = center_aligned_text
        ws.merge_cells(start_row=1, start_column=column_number, end_row=1, end_column=column_number + 1)
        ws.cell(2, column_number, 'Value')
        if normalize:
            ws.cell(2, column_number + 1, 'Frequency')
        else:
            ws.cell(2, column_number + 1, 'Count')
        row_number = 3
        for value, count in value_counts.items():
            ws.cell(row_number, column_number, value if not pd.isna(value) else 'Missing Data')
            ws.cell(row_number, column_number + 1, count)
            row_number += 1
        table_range = (f"{get_column_letter(column_number)}2:"
                       f"{get_column_letter(column_number + 1)}{len(value_counts) + 1}")
        display_name = (f"{table_name_id}"
                        f"{column.replace(' ', '_')}_"
                        f"{'frequencies' if normalize else 'counts'}")
        table = Table(displayName=display_name, ref=table_range)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True
        )
        ws.add_table(table)
        column_number += 3
        pass


def get_cell_number_format(
        value: Any
) -> str:
    """
    Returns a format string that can be passed to cell.number_format in openpyxl.
    :param value: The value for which to generate a format string.
    :return: A format string interpretable by openpyxl.
    """
    # Use default string formatting for strings.
    if isinstance(value, str):
        number_format = '@'

    # Use ISO standard for dates and times
    elif isinstance(value, (datetime.datetime, pd.Timestamp)):
        number_format = 'YYYY-MM-DD HH:MM:SS'

    elif np.issubdtype(type(value), np.number) or isinstance(value, (float, int)):
        # zero is zero
        if value == 0:
            number_format = '0.'
        # comma-separated values for integers.
        elif value % 1 == 0:
            number_format = '#,###,###,###,###,###'
        else:
            # For floats less than zero, use leading zero and up to four decimal places.
            if value < 1:
                number_format = '0.####'
            # For floats greater than zero use comma-separated values and up to four decimal places.
            else:
                number_format = '#,###,###,###,###,###.####'

    # Use default string formatting for booleans.
    elif isinstance(value, (bool, np.bool_)):
        number_format = '@'

    else:
        raise ValueError

    return number_format


def check_for_duplicate_columns(
        df: pd.DataFrame,
        column: str | pd.Series | None = None
) -> list[str] | dict[str, list[str]]:
    """
    Check to see if there are duplicate columns in a dataframe. Behavior depends on the type of argument
    passed to column.
    If column is None, each column of df will be compared to every other column, and a dictionary will be
    returned. Dictionary keys are column names, and values are lists of columns containing identical values.
    A column name may be passed as a string, in which case all other columns will be compared to that
    column, and a list of columns with identical data will be returned.
    If a series is passed, all columns are compared to the series. A list of the names of identical columns
    is returned.
    :param df: The dataframe to analyze.
    :param column: The name of a column in the dataframe to compare to the other columns. A series to
    compare to the columns of the dataframe, or None, in which case each column of the dataframe is
    compared to every other column.
    :return: Either a list of matching columns in the dataframe, or a dictionary whose keys are the names
    of the dataframe's columns, and whose values are lists of the names of columns with identical data.
    """

    # If the name of a column in the dataframe is passed, compare that column to all others:
    if isinstance(column, str):
        if column not in df.columns:
            raise ValueError(f"{column} not found in dataframe.")
        results = [c for c in df.columns if c != column and df[c].equals(df[column])]

    # If a Series is passed, compare it to all columns:
    elif isinstance(column, pd.Series):
        results = [c for c in df.columns if df[c].equals(df[column])]

    # If column is None, compare each column of the dataframe to every other column:
    elif column is None:
        results = {
            c: [c_2 for c_2 in df.columns if c != c_2 and df[c].equals(df[c_2])]
            for c in df.columns
        }

    else:
        raise ValueError()

    return results


def check_for_functionally_duplicate_columns(
        df: pd.DataFrame,
        column: str | pd.Series | None = None,
        exclude_true_duplicates: bool = True
) -> list[str] | dict[str, list[str]]:
    """
    Check to see if there are functionally duplicate columns in a dataframe. Functionally duplicate columns
    are those that have a one-to-one mapping between them. Behavior depends on the type of argument
    passed to column.
    If column is None, each column of df will be compared to every other column, and a dictionary will be
    returned. Dictionary keys are column names, and values are lists of columns containing identical values.
    A column name may be passed as a string, in which case all other columns will be compared to that
    column, and a list of columns with identical data will be returned.
    If a series is passed, all columns are compared to the series. A list of the names of identical columns
    is returned.
    This can be memory intensive due to duplication of the dataframe.
    :param df: The dataframe to analyze.
    :param column: The name of a column in the dataframe to compare to the other columns. A series to
    compare to the columns of the dataframe, or None, in which case each column of the dataframe is
    compared to every other column.
    :param exclude_true_duplicates: If True, Series and columns that are identical will be excluded. Only
    functionally duplicate columns will be returned.
    :return: Either a list of matching columns in the dataframe, or a dictionary whose keys are the names
    of the dataframe's columns, and whose values are lists of the names of columns that match.
    """

    # TODO: Make this less memory-intensive. Currently very fast, but there should be a way to make this both
    #  fast and memory efficient?

    # If the name of a column in the dataframe is passed, compare that column to all others.:
    if isinstance(column, str):
        if column not in df.columns:
            raise ValueError(f"'{column}' not found in DataFrame.")
        functional_duplicates = list()
        for c in df.columns:
            if c != column and not (exclude_true_duplicates and df[c].equals(df[column])):
                pairs = list(set(list(zip(df[c], df[column]))))
                c_vals, c2_vals = zip(*pairs)
                if len(c_vals) == len(set(c_vals)) and len(c2_vals) == len(set(c2_vals)):
                    functional_duplicates.append(c)
        results = functional_duplicates

    # If a Series is passed, compare it to all columns:
    elif isinstance(column, pd.Series):
        if len(column) != df.shape[0]:
            raise ValueError('series length not equal to number of rows in DataFrame.')
        functional_duplicates = list()
        for c in df.columns:
            if not (exclude_true_duplicates and df[c].equals(column)):
                pairs = list(set(list(zip(df[c], column))))
                c_vals, c2_vals = zip(*pairs)
                if len(c_vals) == len(set(c_vals)) and len(c2_vals) == len(set(c2_vals)):
                    functional_duplicates.append(c)
        results = functional_duplicates

    # If column is None, compare each column of the dataframe to every other column:
    elif column is None:
        results = dict()
        for c in df.columns:
            functional_duplicates = list()
            for c2 in df.columns:
                if c != c2 and not (exclude_true_duplicates and df[c].equals(df[c2])):
                    pairs = list(set(list(zip(df[c], df[c2]))))
                    c_vals, c2_vals = zip(*pairs)
                    if len(c_vals) == len(set(c_vals)) and len(c2_vals) == len(set(c2_vals)):
                        functional_duplicates.append(c2)
            results[c] = functional_duplicates

    else:
        raise ValueError()

    return results


def get_dtype_counts(
        data: pd.DataFrame | pd.Series
) -> dict[str, pd.Series] | pd.Series:
    """
    Returns the counts of the 'ideal' data types occurring in a column of data or in a series. Pandas
    defaults to the object dtype when multiple data types are present. The string data type is also a
    relatively recent addition. This method looks at each value in a column or series and determines what
    its optimal data type is, then counts their occurrences.
    If data is a dataframe, then a dictionary of values counts, one for each column, is returned. If data
    is a series, then a series of values counts is returned.
    :param data: Either a dataframe or a series on which to count datatypes.
    :return: Either a dictionary of value count series, or a single value count series, depending on
    whether a dataframe or series is passed to data.
    """

    def classify_type(x):
        # type(x) will classify strings as 'object' by type(x), despite the presence of a superior string
        # dtype in pandas.
        if isinstance(x, str):
            return 'string'
        # NaN values are classified as 'float' by type(x)
        elif np.isnan(x):
            return 'NaN'
        else:
            type(x)

    # If data is a DataFrame:
    if isinstance(data, pd.DataFrame):
        results = dict()
        for column in data.columns:
            types = data[column].apply(classify_type)
            type_counts = types.value_counts()
            type_counts.index = type_counts.index.map(
                lambda x: x if x in ['string', 'NaN'] else x.__name__
            )
            results[column] = type_counts

    # If data is a Series:
    elif isinstance(data, pd.Series):
        types = data.apply(classify_type)
        type_counts = types.value_counts()
        type_counts.index = type_counts.index.map(
            lambda x: x if x in ['string', 'NaN'] else x.__name__
        )
        results = type_counts

    else:
        raise ValueError

    return results


def check_column_order(
        column: pd.Series
) -> str:
    """
    Determines the order of values in a Series.
    :param column: A Series to determine the order of.
    :return: One of 'Ascending', 'Strictly Ascending', 'Descending', 'Strictly Descending', 'Unordered',
    and 'Identical Values'.
    """
    # Drop missing values for the order check
    column = column.dropna()

    # Check for non-comparability (e.g., mixed types)
    try:
        # Check if all values in column are the same:
        if column.nunique() == 1:
            result = "Identical Values"

        # Check for ascending or descending order
        elif all(column.shift(-1)[:-1] > column[:-1]):
            result = "Strictly Ascending"
        elif all(column.shift(1)[:-1] < column[:-1]):
            result = "Strictly Descending"
        elif all(column.shift(-1)[:-1] >= column[:-1]):
            result = "Ascending"
        elif all(column.shift(-1)[:-1] <= column[:-1]):
            result = "Descending"

        # Default to a result of unordered
        else:
            result = "Unordered"

    except ValueError:
        result = "Mixed Data Types"

    return result


def get_fraction_ascending(
        column: pd.Series,
        strict: bool = True
) -> float:
    """
    The fraction of values that are ascending in a Series. A value is classified as ascending if it is
    greater than or equal to the prior value. If strict is True, then a value is classified as ascending only
    if it is greater than the prior value. NaNs are ignored for this calculation. In calculating the 
    fraction, the denominator is the number of non-NaN values minus 1, as the first non-NaN value in the 
    series cannot be classified without a prior value to compare it to.
    :param column: The series containing the data to be analyzed.
    :param strict: If True, values are classified as ascending only if they are greater than the prior 
    non-NaN value. If False, values need only be greater than or equal to the prior non-NaN value.
    :return: The number of values classified as ascending divided by the number of non-NaN values minus one.
    """
    column = column.dropna()
    if strict:
        result = sum(column.shift(-1)[:-1] > column[:-1]) / (column.count() - 1)
    else:
        result = sum(column.shift(-1)[:-1] >= column[:-1]) / (column.count() - 1)
    return result


def get_fraction_descending(
        column: pd.Series,
        strict: bool = True
) -> float:
    """
    The fraction of values that are descending in a Series. A value is classified as descending if it is
    less than or equal to the prior value. If strict is True, then a value is classified as descending only
    if it is less than the prior value. NaNs are ignored for this calculation. In calculating the 
    fraction, the denominator is the number of non-NaN values minus 1, as the first non-NaN value in the 
    series cannot be classified without a prior value to compare it to.
    :param column: The series containing the data to be analyzed.
    :param strict: If True, values are classified as descending only if they are less than the prior 
    non-NaN value. If False, values need only be less than or equal to the prior non-NaN value.
    :return: The number of values classified as descending divided by the number of non-NaN values minus one.
    """
    column = column.dropna()
    if strict:
        result = sum(column.shift(-1)[:-1] < column[:-1]) / (len(column) - 1)
    else:
        result = sum(column.shift(-1)[:-1] <= column[:-1]) / (len(column) - 1)
    return result


def get_column_summary(
        df: pd.DataFrame,
        check_fx_cor: bool = True
) -> dict[str, dict[str, Any]]:
    """
    Generates basic summary data for each column of the DataFrame. Primarily numerical summary data is
    calculated in get_numeric_column_summary().
    :param df: The DataFrame to analyze.
    :param check_fx_cor: Whether to check for functional correlation between columns. Functionally
    correlated columns have 1:1 mappings between values. This can be memory intensive.
    :return: A dictionary whose keys are column names and whose values are sub-dictionaries of summary data
    for the column. Each sub-dictionary contains identical sets of keys.
    """

    # Initialize a dictionary to hold the results
    results = {
        'Data Type': dict(),
        'Unique Values': dict(),
        'Missing Data Points, Count': dict(),
        'Missing Data Points, Fraction': dict(),
        'Most Frequent Value': dict(),
        'Most Frequent Value, Count': dict(),
        'Most Frequent Value, Fraction': dict(),
        'Shannon Entropy': dict(),
        'Maximum Possible Entropy': dict(),
        'Mode Frequency Ratio': dict(),
        'Memory Usage (MB)': dict(),
        'Duplicate Fields': dict()
    }
    if check_fx_cor:
        results['Functionally Duplicate Fields'] = dict()

    # Some calculations only need to be done once, rather than repeated as we iterate through the columns.
    memory_usage = df.memory_usage(deep=True)

    if check_fx_cor:
        functional_correlation_results = check_for_functionally_duplicate_columns(df)

    # Calculate metrics for each column
    for column in df.columns:
        results['Data Type'][column] = str(df[column].dtype)
        results['Unique Values'][column] = df[column].nunique()
        results['Missing Data Points, Count'][column] = df[column].isnull().sum()
        results['Missing Data Points, Fraction'][column] = df[column].isnull().mean()
        if df[column].isna().all():
            results['Most Frequent Value'][column] = 'No Data in Field'
            results['Most Frequent Value, Count'][column] = 'N/A'
            results['Most Frequent Value, Fraction'][column] = 'N/A'
        elif not (df[column].duplicated(keep=False) & ~df[column].isna()).any():
            results['Most Frequent Value'][column] = 'No Duplicate Values'
            results['Most Frequent Value, Count'][column] = 'N/A'
            results['Most Frequent Value, Fraction'][column] = 'N/A'
        else:
            mode_value = df[column].mode()[0]
            mode_value_count = df[column].value_counts()[mode_value]
            results['Most Frequent Value'][column] = mode_value
            results['Most Frequent Value, Count'][column] = mode_value_count
            results['Most Frequent Value, Fraction'][column] = mode_value_count / len(df[column])
        probabilities = df[column].value_counts(normalize=True)
        results['Shannon Entropy'][column] = -np.sum(probabilities * np.log2(probabilities))
        results['Maximum Possible Entropy'][column] = np.log2(len(probabilities))
        max_freq = probabilities.iloc[0]
        second_max_freq = probabilities.iloc[1] if len(probabilities) > 1 else 0
        results['Mode Frequency Ratio'][column] = (
            max_freq / second_max_freq if second_max_freq > 0 else 'N/A'
        )
        results['Memory Usage (MB)'][column] = memory_usage[column] / (1024 ** 2)
        duplicate_columns = [c for c in df.columns if c != column and df[c].equals(df[column])]
        results['Duplicate Fields'][column] = (
            ', '.join(duplicate_columns) if len(duplicate_columns) > 0 else 'None'
        )
        if check_fx_cor:
            results['Functionally Duplicate Fields'][column] = (
                ', '.join(functional_correlation_results[column])
                if len(functional_correlation_results[column]) > 0
                else 'None'
            )

    return results


def get_numeric_column_summary(
        df: pd.DataFrame
) -> dict[str, dict[str, Any]]:
    """
    Generates summary data specifically for numeric columns of the DataFrame. Ignores columns with
    non-numeric data types.
    :param df: The DataFrame to analyze.
    :return: A dictionary whose keys are column names and whose values are sub-dictionaries of summary data
    for the column. Each sub-dictionary contains identical sets of keys.
    """

    # Initialize a dictionary to hold the results
    results = {
        'Mean': dict(),
        'Standard Deviation': dict(),
        'Coefficient of Variation': dict(),
        'Skewness': dict(),
        'Excess Kurtosis': dict(),
        'Minimum': dict(),
        '5th Percentile': dict(),
        '10th Percentile': dict(),
        '25th Percentile': dict(),
        'Median': dict(),
        '75th Percentile': dict(),
        '90th Percentile': dict(),
        '95th Percentile': dict(),
        'Maximum': dict(),
        'Interquartile Range': dict(),
        'Outlier Lower Threshold': dict(),
        'Outlier Upper Threshold': dict(),
        'Outliers, Count': dict(),
        'Outliers, Fraction': dict(),
        'Range': dict(),
        'Zeros, Count': dict(),
        'Zeros, Fraction': dict()
    }

    # Calculate metrics for each column
    for column in df.columns:
        if np.issubdtype(df[column].dtype, np.number) and not df[column].isna().all():
            column_mean = df[column].mean()
            results['Mean'][column] = column_mean
            column_std = df[column].std()
            results['Standard Deviation'][column] = column_std
            results['Coefficient of Variation'][column] = (
                column_std / column_mean if column_mean != 0 else 'Inf'
            )
            results['Skewness'][column] = df[column].skew()
            results['Excess Kurtosis'][column] = df[column].kurt()
            results['Minimum'][column] = df[column].min()
            results['5th Percentile'][column] = df[column].quantile(0.05)
            results['10th Percentile'][column] = df[column].quantile(0.10)
            percentile_25th = df[column].quantile(0.25)
            results['25th Percentile'][column] = percentile_25th
            results['Median'][column] = df[column].median()
            percentile_75th = df[column].quantile(0.75)
            results['75th Percentile'][column] = percentile_75th
            results['90th Percentile'][column] = df[column].quantile(0.90)
            results['95th Percentile'][column] = df[column].quantile(0.95)
            results['Maximum'][column] = df[column].max()
            iqr = percentile_75th - percentile_25th
            results['Interquartile Range'][column] = iqr
            lower_bound = percentile_25th - 1.5 * iqr
            upper_bound = percentile_75th + 1.5 * iqr
            outlier_count = sum((df[column] < lower_bound) | (df[column] > upper_bound))
            results['Outlier Lower Threshold'][column] = lower_bound
            results['Outlier Upper Threshold'][column] = upper_bound
            results['Outliers, Count'][column] = outlier_count
            results['Outliers, Fraction'][column] = outlier_count / len(df[column])
            results['Range'][column] = df[column].max() - df[column].min()
            results['Zeros, Count'][column] = (df[column] == 0).sum()
            results['Zeros, Fraction'][column] = (df[column] == 0).sum() / df.shape[0]

    return results


def get_column_ordering(
        df: pd.DataFrame
) -> dict[str, dict[str, Any]]:
    """
    Generates ordering data for each column of the DataFrame.
    :param df: The DataFrame to analyze.
    :return: A dictionary whose keys are column names and whose values are sub-dictionaries of summary data
    for the column. Each sub-dictionary contains identical sets of keys.
    """

    # Initialize a dictionary to hold the results
    results = {
        'Order': dict(),
        'Fraction Strictly Ascending': dict(),
        'Fraction Ascending': dict(),
        'Fraction Strictly Descending': dict(),
        'Fraction Descending': dict(),
        'Ascending/Descending Ratio': dict(),
        'Log of A/D Ratio': dict()
    }

    # Calculate metrics for each column
    for column in df.columns:
        results['Order'][column] = check_column_order(df[column])
        results['Fraction Strictly Ascending'][column] = get_fraction_ascending(df[column])
        f_ascending = get_fraction_ascending(df[column], strict=False)
        results['Fraction Ascending'][column] = f_ascending
        results['Fraction Strictly Descending'][column] = get_fraction_descending(df[column])
        f_descending = get_fraction_descending(df[column], strict=False)
        results['Fraction Descending'][column] = f_descending
        results['Ascending/Descending Ratio'][column] = (
            f_ascending / f_descending if f_descending != 0 else 'Inf'
        )
        if f_descending == 0:
            log_ratio = 'Inf'
        elif f_ascending == 0:
            log_ratio = '-Inf'
        else:
            log_ratio = np.log(f_ascending / f_descending)
        results['Log of A/D Ratio'][column] = log_ratio

    return results


def get_stacked_umap_decomposition(
        df: pd.DataFrame,
        use_copy_of_df: bool = True,
        umap_transformer_kwargs: dict | None = None,
        random_seed: int | None = None
) -> pd.DataFrame:
    """
    Takes a dataframe, and reduces its row dimension using UMAP (https://umap-learn.readthedocs.io).
    Treats each row as a point in a space of dimension equal to the number of columns in the dataframe,
    and finds an optimal non-linear lower-dimensional manifold on which to project it using fuzzy topology.
    :param df: The dataframe containing the data to reduce the dimension of.
    :param use_copy_of_df: If true, when non-numerical data is encoded the original dataframe columns will
    not be overwritten. May be memory intensive for large dataframes.
    :param umap_transformer_kwargs: Keyword arguments to pass to the UMAP transformer.
    :param random_seed: A random random_seed or None. Note that UMAP can run in parallel on multiple processors,
    but only if no random random_seed is used. Default is None.
    :return: A dataframe containing the reduced data. Uses the same index as the original dataframe.
    """

    # If no umap_transformer_kwargs were passed, set default kwargs.
    # Default is to reduce the dataset to two dimensions.
    if umap_transformer_kwargs is None:
        umap_transformer_kwargs = dict(
            n_components=2
        )

    # Convert non-numeric data to numeric data using scikit-learn's LabelEncoder method.
    # Incorrect ordering of ordinal data has little impact on UMAP's performance, but manual ordering of
    # ordinal data via integer encoding can be managed prior to handing data to this method (Or 1-hot or
    # other categorical data encoding).
    encoder = LabelEncoder()
    if use_copy_of_df:
        df = df.copy()
    for column in df.columns:
        if not pd.api.types.is_numeric_dtype(df[column].dtype):
            df[column] = encoder.fit_transform(df[column])

    # Replace missing values with zeros
    # Crude. Better approaches are recommended before passing a dataframe to this method.
    df.fillna(0, inplace=True)

    # Initialize the umap transformer.
    umap_transformer = UMAP(**umap_transformer_kwargs, random_state=random_seed)

    # Suppress warning generated by fit_transform when used with a random seed.
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message=".*Use no seed for parallelism*.")
        # Decompose the data
        decomposed_data = umap_transformer.fit_transform(df.values)

    # Pack it back into a dataframe, using the original index from the initial dataframe:
    decomposed_df = pd.DataFrame(
        decomposed_data,
        columns=[f'umap_dim_{i}' for i in range(decomposed_data.shape[1])],
        index=df.index
    )

    return decomposed_df


def get_hdbscan_cluster_labels(
        df: pd.DataFrame,
        clustering_kwargs: dict | None = None,
        random_seed: int | None = None
) -> dict[int, list[int]]:
    """
    Applies HDBSCAN (https://hdbscan.readthedocs.io) to classify datapoints in a dataframe into clusters.
    :param df: The dataframe containing the data points.
    :param clustering_kwargs: Keyword arguments to pass to the clustering algorithm.
    :param random_seed: A random seed or None. HDBSCAN is mostly deterministic, and the primary impact of a
    random seed is to break ties when assigning points to clusters, with little impact. Default is None.
    :return: A dictionary. Keys are cluster ids (integers) and values are list of indices associated with
    that id in the dataframe. While indices >= 0 are arbitrary identifiers of little use, "noise" data
    points not placed in clusters are labeled -1, making this useful data to track.
    """

    # If no keyword arguments were passed, initialize default dictionary.
    if clustering_kwargs is None:
        clustering_kwargs = dict()

    # HDBSCAN does not have a built-in random seed argument, but it's behavior can be made reproducible by
    # setting numpy's random seed.
    if random_seed is not None:
        np.random.seed(random_seed)

    # Initialize model and apply to dataframe.
    model = HDBSCAN(**clustering_kwargs)
    clusters = model.fit_predict(df)

    # Construct a dictionary from the results, with cluster ids (integers) as keys and lists of dataframe
    # indices as values.
    unique_values, indices = np.unique(clusters, return_inverse=True)
    cluster_dict = {value: np.where(clusters == value)[0].tolist() for value in unique_values}

    return cluster_dict


def fig_stacked_umap_decomposition_2d(
        df: pd.DataFrame,
        columns: list[str] | None = None,
        use_copy_of_df: bool = True,
        umap_transformer_kwargs: dict | None = None,
        label_map: dict[int, str] | None = None,
        clustering_kwargs: dict | None = None,
        figure_kwargs: dict | None = None,
        fig_filepath: str | Path | None = None,
        random_seed: int | None = None
) -> None:
    """
    Performs a UMAP decomposition on the data in the passed dataframe, treating the rows as data points
    in a space with dimension equal to the number of columns in the dataframe. Then applies HDBSCAN
    clustering to the resulting decomposed dataframe. Finally, creates an interactive html figure from the
    decomposed data, labeled based on the clustering.
    Data points in the resulting figure are labeled with hover-text containing the data in the original
    dataframe passed. Decomposition / clustering is performed on the full dataframe by default, but a list
    of columns can also be passed to limit the amount of data used to those columns. Hover-text still uses
    the original dataset.
    Non-numeric data will be converted prior to decomposition using Scikit-learn's LabelEncoder.
    Keyword arguments can be passed for UMAP, HDBSCAN, and the call to plotly's Figure constructor.
    :param df: The dataframe to analyze.
    :param columns: The columns of the dataframe to include when decomposing and clustering. If None,
    the full dataset will be used.
    :param use_copy_of_df: Whether to duplicate the dataset rather than overwriting it when converting
    categorical data to numerical data.
    :param umap_transformer_kwargs: Keyword arguments to pass to the UMAP transformer.
    :param label_map: A map to use to relabel clusters. If None, clusters will be labeled with integers,
    starting at 0, and unclustered data points will be labeled -1. After a first application of this method
    with a random seed, a label_map can be developed based on assessment of the figure. Map should be a
    dictionary with integer keys and string values. All integer labels must be included in the map.
    :param clustering_kwargs: Keyword arguments to pass to HDBSCAN.
    :param figure_kwargs: Keyword arguments to pass to go.Figure.
    :param random_seed: A random seed for reproducibility. This may affect UMAP performance,
    see get_stacked_umap_decomposition documentation for more details.
    :return: None
    """

    # If no columns were specified for umap decomposition, use all columns.
    if columns is None:
        columns = df.columns

    # Get decomposed data
    decomposed_df = get_stacked_umap_decomposition(
        df[columns], use_copy_of_df,
        umap_transformer_kwargs, random_seed
    )

    # Get clusters
    cluster_dict = get_hdbscan_cluster_labels(decomposed_df, clustering_kwargs)
    clusters = {
        cluster: decomposed_df.values[indices, :]
        for cluster, indices in cluster_dict.items()
    }

    # Build labels for clusters
    labels = df.apply(lambda row: '<br>'.join(f"{col}: {row[col]}" for col in df.columns), axis=1)
    if label_map is None:
        cluster_labels = {
            cluster: labels[indices].tolist()
            for cluster, indices in cluster_dict.items()
        }
    else:
        cluster_labels = {
            label_map[cluster]: labels[indices].tolist()
            for cluster, indices in cluster_dict.items()
        }

    # If figure_kwargs is None, generate default dictionary.
    if figure_kwargs is None:
        figure_kwargs = dict()

    # Generate figure.
    fig = go.Figure(
        data=[
            go.Scattergl(
                x=coordinates[:, 0],
                y=coordinates[:, 1],
                name=f"{cluster_id}",
                mode='markers',
                marker=dict(
                    opacity=0.9
                ),
                text=cluster_labels[cluster_id]
            )
            for cluster_id, coordinates in clusters.items()
        ],
        **figure_kwargs
    )

    # Save figure.
    if fig_filepath is None:
        fig_filepath = './default_umap_fig.html'
    fig.write_html(fig_filepath)


if __name__ == '__main__':
    main()
