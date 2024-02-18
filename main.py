import os
import pickle
import warnings

# Matplotlib will be our primary visualization tool:
from matplotlib import pyplot as plt
from matplotlib.dates import DateFormatter
from matplotlib.ticker import FixedLocator, PercentFormatter, FuncFormatter
# We'll use openpyxl to generate our spreadsheets:
from openpyxl.cell import Cell
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Color, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

# There's an irksome warning that pops up when importing pandas currently to make people aware of a new
# dependency that will be coming in 3.0, so let's suppress that:
with warnings.catch_warnings():
    warnings.simplefilter('ignore')
    import pandas as pd
# Seaborn has some useful matplotlib-derived methods, and it's theming is very nice.
import seaborn as sns

# Import my own toolkit for generating summary analyses of pandas DataFrames:
import toolkit as toolkit

# Set a seaborn theme for our matplotlib figures:
sns.set_theme('paper')


# Normally I use latex for more professional results in my matplotlib figures, but that requires
# installation of components outside of those installed by a requirements.txt file, so I will not do
# that here.
# latex_preamble = r'\setlength{\parindent}{0pt}'
# plt.rcParams.update({
#     "text.usetex": True,
#     "font.family": "Times",
#     "text.latex.preamble": latex_preamble
# })


def main():
    # I'm going to create some directories to contain eda figures and summaries.
    eda_dirpath = './exploratory_data_analysis'
    eda_summary_dirpath = os.path.join(eda_dirpath, 'toolkit_summaries')
    eda_fig_dirpath = os.path.join(eda_dirpath, 'exploratory_figures')
    os.makedirs(eda_summary_dirpath, exist_ok=True)
    os.makedirs(eda_fig_dirpath, exist_ok=True)

    # Also some directories to contain figures and data for the final report.
    results_dirpath = './results'
    results_fig_dirpath = os.path.join(results_dirpath, 'figures')
    xlsx_fig_dirpath = os.path.join(results_dirpath, 'xlsx_figures')
    os.makedirs(results_fig_dirpath, exist_ok=True)
    os.makedirs(xlsx_fig_dirpath, exist_ok=True)

    # Start by loading our datasets:
    session_counts_df = pd.read_csv('./data/DataAnalyst_Ecom_data_sessionCounts.csv')
    adds_to_cart_df = pd.read_csv('./data/DataAnalyst_Ecom_data_addsToCart.csv')

    # Perform EDA and data cleaning on sessions_counts_df
    session_counts_df = explore_and_clean_session_counts_df(
        session_counts_df, eda_summary_dirpath, eda_fig_dirpath
    )

    # Perform EDA and data cleaning on adds_to_cart_df
    adds_to_cart_df = explore_and_clean_adds_to_cart_df(
        adds_to_cart_df, eda_summary_dirpath, eda_fig_dirpath
    )

    # I'm going to generate a directory of figures that are useful given this dataset. It would be my
    # normal practice to provide a client with a set of pdf files of figures, along with a pdf document
    # that included those figures with detailed explanations as part of a narrative.
    generate_figures(session_counts_df, adds_to_cart_df, results_fig_dirpath)

    # Generate the specified xlsx file.
    generate_xlsx_file(session_counts_df, adds_to_cart_df, xlsx_fig_dirpath, results_dirpath)

    pass


def explore_and_clean_session_counts_df(
        df: pd.DataFrame,
        summary_dirpath: str,
        eda_fig_dirpath: str
) -> pd.DataFrame:
    # I like to cache expensive method output. Often I'll use file names constructed from hashes of
    # relevant variable input, but here we'll always be using the same dataframe as input.
    cache_filepath = 'data/cleaned_sessions_df.pickle'
    if os.path.exists(cache_filepath):
        with open(cache_filepath, 'rb') as file:
            # Bad practice to return from multiple locations in a method, I know.
            return pickle.load(file)

    # At first glance, this looks like long form table / tidy data.
    # I'm going to use a tool I built to generate spreadsheets with useful summary data:
    toolkit.build_dataset_summary_wb(
        df, os.path.join(summary_dirpath, 'session_counts_initial_summary.xlsx'),
        include_raw_data=True
    )

    # Looking at the file "session_counts_initial_summary.xlsx" in
    # "exploratory_data_analysis/toolkit_summaries":

    # df has three numerical fields and three categorical fields. It does not appear
    # to have a primary key. There are no missing data points. Lots of zeros, but those appear to be
    # meaningful and to not represent missing data.

    # It looks like we have one row per combination of categorical data (datetime data is ordinal data,
    # so categorical), with associated numerical data for that particular combination. Let's just verify
    # that that's the case.
    duplicate_count = sum(df.duplicated(
        subset=['dim_browser', 'dim_deviceCategory', 'dim_date'], keep=False
    ))

    # duplicate_count is zero, so that confirms our hypothesis. The formatting of this table is often
    # referred to as 'long-form' or 'tidy', with each row containing a unique combination of system
    # property values.

    # Data is initially mostly sorted by dim_date, which appears to be dates that have been mis-coded as
    # strings. If converted to a date-time format, it looks like the dataset will be entirely sorted by
    # date initially. Let's change that column to timestamps:
    df['dim_date'] = pd.to_datetime(df['dim_date'], format='%m/%d/%y')

    # Looking at the dataframe with my IDE, it looks like we have an uninterrupted series of dates in the
    # range [2012-07-01, 2013-06-30]. Let's verify that:
    dttm_first = pd.to_datetime('2012-07-01')
    dttm_last = pd.to_datetime('2013-06-30')
    expected_date_range = pd.date_range(start=dttm_first, end=dttm_last)
    expected_set = set(expected_date_range)
    actual_set = set(df['dim_date'])
    set_size_difference = len(expected_set - actual_set)

    # No dates were missing, so we have uninterrupted data. However, given that this looks like time series
    # data, and given that we are working with increments of days, we should verify that the first and last
    # day of data are complete. We shouldn't assume that recording of data started at or ended at midnight.
    # We can't definitively rule this out, but we can sum the numerical columns for each day and verify
    # that the first and last day sums do not deviate dramatically.
    sub_df = df[['dim_date', 'sessions', 'transactions', 'QTY']]
    grouped_df = sub_df.groupby('dim_date').agg(dict(
        sessions='sum',
        transactions='sum',
        QTY='sum'
    ))
    # Let's look at a toolkit summary of our grouped dataframe:
    toolkit.build_dataset_summary_wb(
        grouped_df.reset_index(),
        os.path.join(summary_dirpath, 'session_counts_grouped_on_dim_date_summary.xlsx'),
        include_raw_data=True
    )

    # Looking at the summary statistics, and at the grouped data, we see that the last day of the period
    # has values very close to the mean. However, we see that the first day of the period has values below
    # the 5th percentile, which is concerning. This is time series data with a granularity of one day,
    # however, so it is worth looking for a day-of-week cycle. Perhaps some days are slower, and when we
    # look at 2012-07-01 in that context, it will not be a probable outlier.

    # Add day of week data to the dataframe
    grouped_df['day_of_week'] = grouped_df.index.day_name()
    numerical_columns = ['sessions', 'transactions', 'QTY']
    # We'll make summary data for each numerical data point broken down by day of week.
    for column in numerical_columns:
        # Create a table with day of week as the columns
        by_day_df = grouped_df.pivot(columns='day_of_week', values=column)
        # We don't care about the index data, so let's drop the mess of na values and make a new dataframe
        by_day_df = pd.concat(
            [by_day_df[c].dropna().reset_index(drop=True) for c in by_day_df.columns], axis=1
        )
        # And let's take a look at a toolkit summary for the new dataframe:
        toolkit.build_dataset_summary_wb(
            by_day_df, os.path.join(
                summary_dirpath, f'session_counts_grouped_on_day_of_week_{column}_summary.xlsx'
            ),
            include_raw_data=True
        )

    # We observe that while there is some variation by day of week, it is not tremendous. There are also a few
    # other, more dramatic outliers, but this one is at the edge of the dataset, which is concerning.
    # However, we also observe that for every other day of the week, we have exactly 52 data points,
    # but for Sunday (2012-07-01 is a Sunday), we have 53, so it is actually convenient to drop this first
    # date, and seems reasonable to do so. (365 % 7 = 1, so expected - except that we were fortunate
    # and the extra day was our potential bad data point at the edge of our dataset).
    dttm_first = pd.to_datetime('2012-07-02')
    dttm_last = pd.to_datetime('2013-06-30')
    df = df[(df['dim_date'] >= dttm_first) & (df['dim_date'] <= dttm_last)]

    # One more thing to look at for the timestamp column quality-wise is to quickly verify that the number
    # of rows for a given date is not hugely variable. This would potentially be an indication of missing
    # data. For instance, if one day had only a single entry, and every other day had 12-18, that might
    # indicate a problem.
    entries_per_day_df = df['dim_date'].value_counts().value_counts().sort_index()

    # This spread looks a little top-heavy, but there are no terribly small values that would suggest data
    # truncation. I expect that this top-heaviness is due to the highly skewed mess that is the
    # 'dim_browser' column, but first lets look at 'dim_deviceCategory'.

    # 'dim_deviceCategory' consists of only three values, 'mobile', 'desktop', and 'tablet'. These are all
    # well represented, which is good. If one was only present for less than ten out of 7,700 rows of data,
    # I would be concerned that it was bad data. It also passes the sanity test. These are descriptions of
    # the three fundamental types of computing devices on which we access the internet.
    # I am guessing that 'desktop' here covers both laptops and actual desktops. If this were a real-life
    # data set that would be a question I would take back to the customer.
    # I will just capitalize these entries to facilitate generating quality figures later on.
    df = df.copy()  # I subsetted the df earlier...
    df['dim_deviceCategory'] = df['dim_deviceCategory'].str.capitalize()

    # 'dim_browser' is a mess. Let's get counts:
    dim_browser_value_counts_df = df['dim_browser'].value_counts().sort_values(ascending=False)

    # So most importantly - super entertaining that there's a "Nintendo Browser" in here. This is kind of a
    # mess though, but we have two things going for us. First, a relatively few categories are dominant.
    # Secondly, there are only 57 unique values. So it's reasonable to just create a manual map to the
    # values we want. If a little tiresome.

    # From a statistical standpoint, this is extremely right-skewed data, which indicates a high
    # possibility that the extremely rare categories are bad data - especially those that only occur once.
    # Also, it means that interventions addressing a relatively small fraction of the categories will
    # affect the vast majority of events. So it's reasonable to focus on the majority events. We won't
    # throw out the rare events though - we'll just aggregate them together into single category so we can
    # study their properties, and still represent them in the dataset.

    # How we aggregate these values depends on how we are applying the data, and I would be consulting on
    # my client on this. For now, I'm going to do a limited amount of cleanup, including the creation of an
    # 'Unknown' category. I was also going to do an 'Other' category, but there is a lot of weirdness in
    # the rare categories that leads me to lump it all into Unknown.
    # First though, I'll print code to the console that I can copy and paste back to minimize the amount of
    # typing I have to do. Note that I'm sorting it by value counts before printing it.
    printed_dict_values = [
        f"'{value}': '{value}'"
        for value in df['dim_browser'].value_counts().sort_values(ascending=False).index
    ]
    print('{\n\t\t' + ',\n\t\t'.join(printed_dict_values) + '\t\n\t}')
    # Now I'll paste that code below and then zip through and edit it quickly to create our map:
    browser_map = {
        'Chrome': 'Chrome',
        'Internet Explorer': 'Internet Explorer',
        'Safari': 'Safari',
        'Edge': 'Edge',
        'Firefox': 'Firefox',
        'Safari (in-app)': 'Safari (in-app)',
        'Opera': 'Opera',
        'Android Webview': 'Android',  # Change
        'Samsung Internet': 'Samsung Internet',
        'Amazon Silk': 'Amazon Silk',
        'error': 'Unknown',  # Change
        'Android Browser': 'Android',  # Change
        'BlackBerry': 'BlackBerry',
        'SeaMonkey': 'SeaMonkey',
        'Opera Mini': 'Opera Mini',
        'UC Browser': 'UC Browser',
        'Mozilla': 'Mozilla',
        'Maxthon': 'Maxthon',
        'YaBrowser': 'YaBrowser',
        'Puffin': 'Puffin',
        '(not set)': 'Unknown',  # Change
        'Mozilla Compatible Agent': 'Mozilla Compatible Agent',
        'osee2unifiedRelease': 'Unknown',  # Change
        'Coc Coc': 'Coc Coc',  # A Vietnamese browser! That's kind of cool.
        'Iron': 'Iron',
        # There is a lot of junk below here and very small value counts.
        'BrowserNG': 'Unknown',
        'DESKTOP': 'Unknown',
        'Truefitbot': 'Unknown',
        'DDG-Android-3.1.1': 'Unknown',
        'MRCHROME': 'UnknownE',
        'NokiaC7-00': 'Unknown',
        'NokiaE52-1': 'Unknown',
        'YelpWebView': 'Unknown',
        'Seznam': 'Unknown',
        'IE with Chrome Frame': 'Unknown',
        'TimesTablet': 'Unknown',
        'LG-C410': 'Unknown',
        'NetFront': 'Unknown',
        'DDG-Android-3.0.14': 'Unknown',
        'Apple-iPhone7C2': 'Unknown',
        'Nokia Browser': 'Unknown',
        'Amazon.com': 'Unknown',
        'SonyEricssonK700c': 'Unknown',
        'DDG-Android-3.0.11': 'Unknown1',
        'Mobile': 'Unknown',
        'NetNewsWire Browser': 'Unknown',
        'HubSpot inbound link reporting check': 'Unknown',
        'DDG-Android-3.0.17': 'Unknown',
        'X-WebBrowser': 'Unknown',
        'Nintendo Browser': 'Unknown',
        'turnaround': 'Unknown',
        'anonymous': 'Unknown',
        'Playstation 3': 'Unknown',
        'Job Search': 'Unknown',
        'Python-urllib': 'Unknown',
        'FeeddlerPro': 'Unknown',
        'Chromeless 1.2.0': 'Unknown'
    }

    # Apply the map.
    df['dim_browser'] = df['dim_browser'].apply(lambda x: browser_map[x])

    # Now I need to combine some of my numeric columns, as I have duplicate categorical rows.
    df = df.groupby(
        ['dim_browser', 'dim_deviceCategory', 'dim_date'], as_index=False
    ).sum()

    # This reduces the size of our dataframe from 7,733 entries to 7,383 entries.

    # Now we need to take a look at our numeric data. I'm going to generate histograms and summary
    # worksheets for each column for total daily values, grouped across all categories.
    daily_sums_df = df.groupby('dim_date', as_index=False).sum()
    daily_sums_df = daily_sums_df.drop(['dim_browser', 'dim_deviceCategory'], axis=1)
    toolkit.build_dataset_summary_wb(
        daily_sums_df, os.path.join(summary_dirpath, f'session_counts_daily_sums_summary.xlsx'),
        include_raw_data=True
    )
    for column in ['sessions', 'transactions', 'QTY']:
        fig, ax = plt.subplots()  # type: plt.Figure, plt.Axes
        sns.histplot(daily_sums_df[column], ax=ax)
        ax.set_title(
            f"{'QTY' if column == 'QTY' else column.capitalize()} Distribution", fontsize=20
        )
        ax.set_xlabel(f"{'QTY' if column == 'QTY' else column.capitalize()}", fontsize=14)
        ax.set_ylabel('Adds To Cart (Thousands)', fontsize=14)
        fig.savefig(os.path.join(
            eda_fig_dirpath, f"{'QTY' if column == 'QTY' else column.capitalize()} Distribution")
        )

    # All three sets of data have fairly positive skews, with some outliers via the IQR*1.5 rule. They are
    # also all strongly correlated, which makes sense (This is also in my summary workbooks).
    # I would expect increased sessions to be associated with increased transactions and purchase
    # quantities. When sorted by date, the large outliers all seem to be in April, May, and June. Perhaps
    # there was a sale of some sort during that period? I don't see any good reason to exclude the
    # potential outliers in the numerical data. The fact that all values correlate is reassuring. It's less
    # likely that an error in data collection would have affected all three types of data - though still
    # possible.
    # ***If I had access to the client I would confirm this decision with them.***

    # From a sanity standpoint, we also expect that our sessions values should be greater than or equal to
    # our transactions value, and our QTY values should be greater than or equal to our
    # transactions values. Let's confirm this.
    bad_row_df = df[df['sessions'] < df['transactions']]

    # Hmmmm... We have 5 rows where there are more transactions than sessions. I suppose it's possible
    # that, however this data is generated, a user populated a cart more than once during a session?
    # However, there are also two rows with '0' for sessions. This is a tiny portion of the overall data,
    # and the browsers used are moderately exotic, so let's drop these rows
    df = df[df['sessions'] >= df['transactions']]

    # Now the quantity / transactions relationship:
    bad_row_df = df[df['transactions'] > df['QTY']]

    # Well, that's odd. There are 576 records where there are more transactions than quantity. ***At this
    # point I would go to the client and ask questions about how the data was collected.*** As is,
    # this is a very large chunk of data. These are not exotic browsers, and there are large numbers of
    # sessions in many of these records. I won't discard these rows for purposes of this exercise.

    # Another sanity check is there should be no records with zero sessions. If every date had a row for
    # every combination of browser and device category, this would make sense, but we have already observed
    # that the number of entries with a given date is variable. Let's see if there are any rows with zero
    # sessions left at this point.
    zero_transactions_df = df[df['sessions'] == 0]

    # Wow. 500 rows. For the most part, these rows also have zero values for transactions and QTY,
    # but there are three with positive values for quantity. Again, this would have me going back to the
    # client with questions about how the data was generated. For now though we'll just drop all columns
    # where sessions is 0. This is also helpful for calculating our 'ECR' later, where division by zero
    # would otherwise require special treatment.
    df = df[df['sessions'] != 0]

    # The last thing I will do is fix the terrible column names.
    df.columns = ['browser', 'device_category', 'date', 'sessions', 'transactions', 'quantity']

    # And generate a summary workbook for the cleaned dataset
    toolkit.build_dataset_summary_wb(
        df, os.path.join(
            summary_dirpath, f'cleaned_session_counts_summary.xlsx'
        ),
        include_raw_data=True
    )

    # There are a lot of figures we will want to make to further analyze the data, but I'm going to include
    # those in my results, so I guess I'll do them in other methods? This is all a little artificial as EDA
    # and data analysis in general is a cyclical and messy process, but I'm trying to share and display it
    # in a linear fashion for this demo...

    # Let's cache our cleaned dataframe:
    with open(cache_filepath, 'wb') as file:
        pickle.dump(df, file)

    return df


def explore_and_clean_adds_to_cart_df(
        df: pd.DataFrame,
        summary_dirpath: str,
        fig_dirpath: str
) -> pd.DataFrame:
    # I like to cache expensive method output. Often I'll use file names constructed from hashes of
    # relevant variable input, but here we'll always be using the same dataframe as input.
    cache_filepath = 'data/cleaned_adds_df.pickle'
    if os.path.exists(cache_filepath):
        with open(cache_filepath, 'rb') as file:
            # Bad practice to return from multiple locations in a method, I know.
            return pickle.load(file)

    # Let's generate a summary workbook...
    toolkit.build_dataset_summary_wb(
        df, os.path.join(summary_dirpath, 'adds_to_cart_initial_summary.xlsx'),
        include_raw_data=True
    )

    # The adds_to_cart data is a very small dataset - just 12 monthly values and a single piece of
    # numerical data for each month. There are no missing data points. Let's switch out the 'dim_month' and
    # 'dim_year' columns for a proper timestamp, and drop them.
    df['month'] = df.apply(
        lambda row: pd.Timestamp(year=row['dim_year'], month=row['dim_month'], day=1), axis=1
    )
    df = df[['month', 'addsToCart']]

    # Not much to do here in terms of exploration and cleaning. We have no gaps in our times series,
    # with deltas of one month. No outliers by the conservative 1.5*IQR rule, though we only have 12 data
    # points anyway.

    # The only EDA plot to generate here would be, potentially, a basic time series:
    fig, ax = plt.subplots()  # type: plt.Figure, plt.Axes
    sns.lineplot(df, x='month', y='addsToCart', ax=ax)
    fig.set_size_inches(11, 8.5)
    fig.set_dpi(300)
    ax.set_title('Adds to Cart vs Month', fontsize=20)
    ax.set_xlabel('Month', fontsize=14)
    ax.set_ylabel('Adds To Cart (Thousands)', fontsize=14)
    ax.xaxis.set_major_formatter(DateFormatter('%b %Y'))
    tick_locations = ax.get_yticks()
    ax.yaxis.set_major_locator(FixedLocator(tick_locations))
    ax.set_yticklabels(
        [int(label / 1000) for label in tick_locations]
    )
    ax.set_ylim(80_000, 240_000)
    fig.savefig(os.path.join(fig_dirpath, 'Adds to Cart vs Month.pdf'))

    # There is potentially an overall downward trend here, but there is no consistent direction of change
    # from month to month, and we have very few data points.

    # Last thing I'll do is change up the column name to something more pythonic
    df.columns = ['month', 'adds_to_cart']

    # And generate a summary workbook for the cleaned dataset
    toolkit.build_dataset_summary_wb(
        df, os.path.join(
            summary_dirpath, f'cleaned_adds_to_cart_summary.xlsx'
        ),
        include_raw_data=True
    )

    # Like I said in the previous EDA method:
    # There are a lot of figures we will want to make to further analyze the data, but I'm going to include
    # those in my results, so I guess I'll do them in other methods? This is all a little artificial as EDA
    # and data analysis in general is a cyclical and messy process, but I'm trying to share and display it
    # in a linear fashion for this demo...

    # Let's cache our cleaned dataframe:
    with open(cache_filepath, 'wb') as file:
        pickle.dump(df, file)

    return df


def generate_figures(
        sessions_df: pd.DataFrame,
        adds_df: pd.DataFrame,
        fig_dirpath: str
) -> None:
    # As previously mentioned, I would usually use latex to annotate figures - it looks much more
    # professional, is more flexible, and allows for inclusion of equations. However, as it requires
    # installation of system packages that cannot be managed with a requirements file, I will not use latex
    # here.

    # I'm going to use arbitrary figure numbers here to organize the folder. These would normally correlate
    # to a written report. I would also typically build a subdirectory structure for a large project,
    # and put more time into figure labeling. Also, the order of the numbering doesn't represent the order
    # in which I would tell the story of the data, necessarily.

    # Let's generate histograms and box plots for our numeric data in sessions_df:
    df = sessions_df.groupby(['date']).sum()[['sessions', 'transactions', 'quantity']]
    # We'll need to calculate an ecr value for this dataframe based on our grouping:
    df['ecr'] = df['transactions'] / df['sessions']

    columns = ['sessions', 'transactions', 'quantity', 'ecr']
    names = ['Sessions', 'Transactions', 'Quantity', 'ECR']

    # First, histograms for each without aggregation:
    for i, (column, name) in enumerate(zip(columns, names)):
        fig, ax = plt.subplots()  # type: plt.Figure, plt.Axes
        sns.histplot(df[column], ax=ax)
        fig.set_size_inches(11, 8.5)
        fig.set_dpi(300)
        title = f"Fig 1{['a', 'b', 'c', 'd'][i]}: Histogram of {name} Distribution"
        ax.set_title(title, fontsize=20)
        ax.set_xlabel(name, fontsize=14)
        ax.set_ylabel("Count", fontsize=14)
        fig.savefig(os.path.join(fig_dirpath, f"{title.replace(':', '')}.pdf"))
        plt.close()

    # Now boxplots for each without aggregation. These are single boxplots, so let's put them all on the
    # same figure, with a subplot for each:
    fig, axs = plt.subplots(nrows=1, ncols=4)  # type: plt.Figure, plt.Axes
    fig.set_size_inches(11, 8.5)
    fig.set_dpi(300)
    fig.subplots_adjust(wspace=0.5)
    title = f"Fig 2: Box Plots of Numerical Sessions Data Distribution"
    fig.suptitle(title, fontsize=20)
    for i, (column, name) in enumerate(zip(columns, names)):
        ax = axs[i]
        sns.boxplot(y=column, data=df, ax=ax, showfliers=False)
        ax.scatter(x=0, y=df[column].mean(), color='black', s=100)
        ax.set_xlabel(name, fontsize=14)
        ax.set_ylabel("", fontsize=14)
    fig.savefig(os.path.join(fig_dirpath, f"{title.replace(':', '')}.pdf"))
    plt.close()

    # Now let's do the same, but break things down by device category:
    df = sessions_df.groupby(['date', 'device_category'], as_index=False).sum().drop(['browser'], axis=1)
    for i, (column, name) in enumerate(zip(columns, names)):
        fig, ax = plt.subplots()  # type: plt.Figure, plt.Axes
        sns.boxplot(x='device_category', y='sessions', data=df, ax=ax, showfliers=False)
        fig.set_size_inches(11, 8.5)
        fig.set_dpi(300)
        title = f"Fig 3{['a', 'b', 'c', 'd'][i]}: Box Plots of Distribution of {name} by Device Category"
        ax.set_title(title, fontsize=20)
        ax.set_xlabel("Device Category", fontsize=14)
        ax.set_ylabel(f"{name}", fontsize=14)
        fig.savefig(os.path.join(fig_dirpath, f"{title.replace(':', '')}.pdf"))
        plt.close()

    # Now figures for these values by day of week:
    # First, without breaking it down by device category:
    df = sessions_df.groupby(['date'], as_index=False).sum()
    df['ecr'] = df['transactions'] / df['sessions']

    # We need an ordered categorical data type for the days of the week:
    day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    df['day_of_week'] = pd.Categorical(df['date'].dt.day_name(), categories=day_order, ordered=True)

    for i, (column, name) in enumerate(zip(columns, names)):
        fig, ax = plt.subplots()  # type: plt.Figure, plt.Axes
        sns.boxplot(x='day_of_week', y=column, data=df, ax=ax, showfliers=False)
        fig.set_size_inches(11, 8.5)
        fig.set_dpi(300)
        title = f"Fig 4{['a', 'b', 'c', 'd'][i]}: Distribution of {name} by Day of Week"
        ax.set_title(title, fontsize=20)
        ax.set_ylabel(f"{name}", fontsize=14)
        fig.savefig(os.path.join(fig_dirpath, f"{title.replace(':', '')}.pdf"))
        plt.close()

    # Now, broken down by device category:
    df = sessions_df.groupby(['date', 'device_category'], as_index=False).sum()
    df['ecr'] = df['transactions'] / df['sessions']

    # We again need an ordered categorical data type for the days of the week:
    day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    df['day_of_week'] = pd.Categorical(df['date'].dt.day_name(), categories=day_order, ordered=True)

    for i, (column, name) in enumerate(zip(columns, names)):
        fig, ax = plt.subplots()  # type: plt.Figure, plt.Axes
        sns.boxplot(x='day_of_week', y=column, hue='device_category', data=df, ax=ax, showfliers=False)
        fig.set_size_inches(11, 8.5)
        fig.set_dpi(300)
        title = f"Fig 5{['a', 'b', 'c', 'd'][i]}: Distribution of {name} by Day of Week and Device Category"
        ax.set_title(title, fontsize=20)
        ax.set_ylabel(f"{name}", fontsize=14)
        fig.savefig(os.path.join(fig_dirpath, f"{title.replace(':', '')}.pdf"))
        plt.close()

    # Those multi-component boxplots contain a lot of information, to the point where it's hard to see what
    # the trends are. Let's do the means with line-plots. Medians might be more representative of a
    # 'typical' value for these distributions given how skewed the distributions are, but means correlate
    # directly with totals, and totals here correlate with revenue which I imagine is what we're after.
    for i, (column, name) in enumerate(zip(columns, names)):
        fig, ax = plt.subplots()  # type: plt.Figure, plt.Axes
        sns.lineplot(
            x='day_of_week', y=column, hue='device_category', data=df,
            estimator='mean', errorbar=None, ax=ax,
            style='device_category', markers={'Tablet': 'o', 'Desktop': 's', 'Mobile': 'D'}
        )
        fig.set_size_inches(11, 8.5)
        fig.set_dpi(300)
        title = f"Fig 6{['a', 'b', 'c', 'd'][i]}: Mean {name} by Day of Week and Device Category"
        ax.set_title(title, fontsize=20)
        ax.set_ylabel(f"{name}", fontsize=14)
        fig.savefig(os.path.join(fig_dirpath, f"{title.replace(':', '')}.pdf"))
        plt.close()

    # Let's do those line plots again, but with 95% bootstrapped confidence intervals:
    for i, (column, name) in enumerate(zip(columns, names)):
        fig, ax = plt.subplots()  # type: plt.Figure, plt.Axes
        sns.lineplot(
            x='day_of_week', y=column, hue='device_category', data=df,
            estimator='mean', errorbar=('ci', 95), ax=ax,
            style='device_category', markers={'Tablet': 'o', 'Desktop': 's', 'Mobile': 'D'}
        )
        fig.set_size_inches(11, 8.5)
        fig.set_dpi(300)
        title = (f"Fig 7{['a', 'b', 'c', 'd'][i]}: "
                 f"Mean {name} by Day of Week and Device Category\n With 95% Confidence Intervals")
        ax.set_title(title, fontsize=20)
        ax.set_ylabel(f"{name}", fontsize=14)
        fig.savefig(os.path.join(fig_dirpath, f"{title.replace(':', '')}.pdf"))
        plt.close()

    # We should also just generate some simple bar plots of totals by group and day of week:
    df = sessions_df.drop(['date'], axis=1).groupby(['device_category'], as_index=False).sum()
    for i, (column, name) in enumerate(zip(columns[:3], names[:3])):
        fig, ax = plt.subplots()  # type: plt.Figure, plt.Axes
        sns.barplot(x='device_category', y=column, data=df, ax=ax)
        fig.set_size_inches(11, 8.5)
        fig.set_dpi(300)
        title = f"Fig 8{['a', 'b', 'c', 'd'][i]}: Total {name} by Device Category"
        ax.set_title(title, fontsize=20)
        ax.set_ylabel(f"{name}", fontsize=14)
        fig.savefig(os.path.join(fig_dirpath, f"{title.replace(':', '')}.pdf"))
        plt.close()

    # That's all the most important stuff. Keep in mind that I consider these crude, hastily generated
    # figures. Without a client to talk to, better domain knowledge, and specific objectives, this is all
    # pretty much shooting into the void.


def generate_xlsx_file(
        sessions_df: pd.DataFrame,
        adds_df: pd.DataFrame,
        xlsx_figs_dirpath: str,
        results_dirpath: str
) -> None:
    # Create a workbook object and get the first worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'Aggregate Data'
    populate_ws_aggregate_data(sessions_df, ws)
    ws = wb.create_sheet('Month over Month')
    populate_ws_month_over_month(sessions_df, adds_df, ws)
    ws = wb.create_sheet('Day of Week')
    populate_ws_day_of_week(sessions_df, xlsx_figs_dirpath, ws)
    wb.save(os.path.join(results_dirpath, 'reference_tables.xlsx'))
    pass


# Quick utility formatting function for numerical values:
def set_cell_number_format(
        cell: Cell,
        value: int | float | None = None,
        add_leading_plus: bool = False,
        dec_places: int | None = None
) -> None:
    if value is None:
        value = cell.value
    if dec_places is None:
        dec_places = 4
    if isinstance(value, int):
        number_format = '###,###,###,###'
    elif isinstance(value, float):
        decimal_zeros = '0' * dec_places
        number_format = f'0.{decimal_zeros}'
    else:
        return None  # Yes, sloppy, but it's just a quick convenience function.
    if add_leading_plus:
        number_format = f"+{number_format};-{number_format};{number_format}"
    cell.number_format = number_format


def populate_ws_aggregate_data(
        sessions_df: pd.DataFrame,
        ws: Worksheet
) -> None:
    # Not using 'browser'
    df = sessions_df.drop('browser', axis=1)

    # There are fancier things you can do, but it's easiest to just set the day for all my timestamps to be
    # the same, then aggregate on that column. It's also very fast.
    df = df.copy()
    df['date'] = df['date'].apply(lambda dt: dt.replace(day=1))
    df = df.groupby(['date', 'device_category'], as_index=False).sum()

    # I've confirmed that there are no zero values in the 'sessions' column, so no divide by zero problem
    # to worry about here. If I was designing code that would likely see reuse I would definitely address
    # this.
    df['ecr'] = df['transactions'] / df['sessions']

    # We should be sorted by date, but let's make double sure
    df.sort_values(by=['date'], inplace=True)

    # Column titles in bold
    columns = ['Month', 'Device', 'Sessions', 'Transactions', 'QTY', 'ECR']
    bold_font = Font(bold=True)
    for column_number, column in enumerate(columns, start=1):
        cell = ws.cell(1, column_number, column)
        cell.font = bold_font

    # Now the dates, with some formatting
    for row_number, value in enumerate(df['date'], start=2):
        ws.cell(row_number, 1, value.strftime('%Y-%m'))

    # Not going to assume the columns are correctly ordered, though they should be.
    columns = ['device_category', 'sessions', 'transactions', 'quantity', 'ecr']
    for column_number, column in enumerate(columns, start=2):
        for row_number, value in enumerate(df[column], start=2):
            cell = ws.cell(row_number, column_number, value)
            set_cell_number_format(cell, value)

    # Turn the data into a table
    table_range = f"A1:{get_column_letter(df.shape[1])}{df.shape[0] + 1}"
    table = Table(displayName='ws_1', ref=table_range)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True
    )
    ws.add_table(table)

    # Freeze the top row and adjust the column widths
    ws.freeze_panes = 'A2'
    for column_number in range(1, 7):
        ws.column_dimensions[get_column_letter(column_number)].width = 12


def populate_ws_month_over_month(
        sessions_df: pd.DataFrame,
        adds_df: pd.DataFrame,
        ws: Worksheet
) -> None:
    # Not using 'browser' or 'device_category'
    df = sessions_df.drop(['browser', 'device_category'], axis=1)

    # There are fancier things you can do, but it's easiest to just set the day for all my timestamps to be
    # the same, then aggregate on that column. It's also very fast.
    df = df.copy()
    df['date'] = df['date'].apply(lambda dt: dt.replace(day=1))
    df = df.groupby(['date'], as_index=False).sum()
    df['ecr'] = df['transactions'] / df['sessions']
    # Doesn't matter 'how' we join, since we've ensured that our primary key column in each dataframe is
    # identical.
    df = pd.merge(df, adds_df, left_on='date', right_on='month', how='inner')
    # This will retain both primary key columns, so let's drop one...
    df = df.drop('date', axis=1)
    # Let's reverse the sort, and make the most recent data be on the left of the worksheet,
    # and more accessible.
    df.sort_values(by='month', inplace=True, ascending=False)

    # Column headers
    column_headers = [month.strftime('%Y-%m') for month in df['month']]
    bold_font = Font(bold=True)
    for column_number, header in enumerate(column_headers, start=2):
        cell = ws.cell(1, column_number, header)
        cell.font = bold_font
        ws.column_dimensions[get_column_letter(column_number)].width = 15

    # The data
    data_columns = ['transactions', 'sessions', 'ecr', 'quantity', 'adds_to_cart']
    section_titles = ['Transactions', 'Sessions', 'ECR', 'Quantity', 'Adds To Cart']

    # Quick utility function to set cell text color based on value
    green_font = Font(color=Color("28803f"))
    red_font = Font(color=Color("a83232"))

    def set_cell_text_color(
            local_cell: Cell,
            local_value: int | float
    ) -> None:
        if local_value > 0:
            local_cell.font = green_font
        elif local_value < 0:
            local_cell.font = red_font

    # Let's define some PatternFills for visibility:
    light_grey_fill = PatternFill(start_color='B9B9B9', end_color='B9B9B9', fill_type='solid')
    dark_grey_fill = PatternFill(start_color='797A7A', end_color='797A7A', fill_type='solid')

    row_number = 2
    for section, (column, title) in enumerate(zip(data_columns, section_titles)):

        # Section title
        cell = ws.cell(row_number, 1, title)
        cell.font = bold_font
        # Add our pattern fill
        for pattern_column_number in range(1, df.shape[0] + 2):
            ws.cell(row_number, pattern_column_number).fill = light_grey_fill
        row_number += 1

        # Previous month data
        ws.cell(row_number, 1, 'Previous Month')
        for column_number, value in enumerate(df[column][1:], start=2):
            cell = ws.cell(row_number, column_number, value)
            set_cell_number_format(cell, value)
        row_number += 1

        # Current month data
        ws.cell(row_number, 1, 'Current Month')
        for column_number, value in enumerate(df[column], start=2):
            cell = ws.cell(row_number, column_number, value)
            set_cell_number_format(cell, value)
        row_number += 1

        # Absolute Difference
        # ws.cell(row_number, 1, 'Absolute Difference')
        for column_number, (current_value, previous_value) in enumerate(
                zip(df[column], df[column][1:]), start=2
        ):
            value = current_value - previous_value
            cell = ws.cell(row_number, column_number, value)
            set_cell_number_format(cell, value, True)
            set_cell_text_color(cell, value)
        row_number += 1

        # Relative Difference
        # ws.cell(row_number, 1, 'Relative Difference')
        for column_number, (current_value, previous_value) in enumerate(
                zip(df[column], df[column][1:]), start=2
        ):
            value = (current_value - previous_value) / previous_value
            cell = ws.cell(row_number, column_number, value)
            cell.number_format = '+0.0%;-0.0%;0.0'
            set_cell_text_color(cell, value)
        row_number += 1

    # It makes sense to break this data down by device category, so let's put that in blocks below the
    # overall month over month. Really I should have thought of this before and wrapped everything above in
    # a loop, so now I'm going to be a little lazy and copy and paste that big block from above into a
    # three-step loop. Not optimal for reusability I admit.

    for device_category in ['Desktop', 'Mobile', 'Tablet']:

        # Filter on device_category
        df = sessions_df[sessions_df['device_category'] == device_category]

        # Not using 'browser' or 'device_category'
        df = df.drop(['browser', 'device_category'], axis=1)

        # There are fancier things you can do, but it's easiest to just set the day for all my timestamps to
        # be the same, then aggregate on that column. It's also very fast.
        df = df.copy()
        df['date'] = df['date'].apply(lambda dt: dt.replace(day=1))
        df = df.groupby(['date'], as_index=False).sum()
        df['ecr'] = df['transactions'] / df['sessions']

        # We don't have a breakdown for Adds to Cart by device category, so we'll skip the join.

        # The data --- removed Add to Cart
        data_columns = ['transactions', 'sessions', 'ecr', 'quantity']
        section_titles = ['Transactions', 'Sessions', 'ECR', 'Quantity']

        # We'll just increment this to add space between blocks, and then add a big header
        row_number += 3
        cell = ws.cell(row_number, 1, f'{device_category} Only')
        cell.font = bold_font
        for pattern_column_number in range(1, df.shape[0] + 2):
            ws.cell(row_number, pattern_column_number).fill = dark_grey_fill
        row_number += 1

        # And do our same big data loop again...
        for section, (column, title) in enumerate(zip(data_columns, section_titles)):

            # Section title
            cell = ws.cell(row_number, 1, title)
            cell.font = bold_font
            # Add our pattern fill
            for pattern_column_number in range(1, df.shape[0] + 2):
                ws.cell(row_number, pattern_column_number).fill = light_grey_fill
            row_number += 1

            # Previous month data
            ws.cell(row_number, 1, 'Previous Month')
            for column_number, value in enumerate(df[column][1:], start=2):
                cell = ws.cell(row_number, column_number, value)
                set_cell_number_format(cell, value)
            row_number += 1

            # Current month data
            ws.cell(row_number, 1, 'Current Month')
            for column_number, value in enumerate(df[column], start=2):
                cell = ws.cell(row_number, column_number, value)
                set_cell_number_format(cell, value)
            row_number += 1

            # Absolute Difference
            # ws.cell(row_number, 1, 'Absolute Difference')
            for column_number, (current_value, previous_value) in enumerate(
                    zip(df[column], df[column][1:]), start=2
            ):
                value = current_value - previous_value
                cell = ws.cell(row_number, column_number, value)
                set_cell_number_format(cell, value, True)
                set_cell_text_color(cell, value)
            row_number += 1

            # Relative Difference
            # ws.cell(row_number, 1, 'Relative Difference')
            for column_number, (current_value, previous_value) in enumerate(
                    zip(df[column], df[column][1:]), start=2
            ):
                value = (current_value - previous_value) / previous_value
                cell = ws.cell(row_number, column_number, value)
                cell.number_format = '+0.0%;-0.0%;0.0'
                set_cell_text_color(cell, value)
            row_number += 1

    # Freeze the first column and row and set first column width:
    ws.freeze_panes = 'B2'
    ws.column_dimensions['A'].width = 20


def populate_ws_day_of_week(
        sessions_df: pd.DataFrame,
        xlsx_figs_dirpath: str,
        ws: Worksheet
) -> None:
    # Drop the browser column:
    df = sessions_df.copy()
    df.drop(['browser'], axis=1, inplace=True)

    # Define a list to order our device_categories
    device_categories = [None, 'Desktop', 'Mobile', 'Tablet']

    # Define a list to order our days of the week and to use as column headers
    days_of_week = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

    # Column Headers
    bold_font = Font(bold=True)
    for column_number, header in enumerate(days_of_week, start=2):
        cell = ws.cell(1, column_number, header)
        cell.font = bold_font
        ws.column_dimensions[get_column_letter(column_number)].width = 15

    # Create a list of the four dataframes we will iterate through
    category_dfs = (
        # First a dataframe that combines all device categories
            [df.drop(['device_category'], axis=1).groupby('date', as_index=False).sum()] +
            # Then one for each device category, in the order specified in our device_categories list.
            [
                df[df['device_category'] == device_category].groupby('date', as_index=False).sum()
                for device_category in device_categories[1:]
            ]
    )
    # Add day of week values
    for category_df in category_dfs:
        category_df['day_of_week'] = category_df['date'].dt.day_name()

    # Let's define some PatternFills for visibility:
    light_grey_fill = PatternFill(start_color='B9B9B9', end_color='B9B9B9', fill_type='solid')
    dark_grey_fill = PatternFill(start_color='797A7A', end_color='797A7A', fill_type='solid')

    # Iterate through device categories
    row_number = 2
    for device_category, cat_df in zip(device_categories, category_dfs):

        # Section title if it's not the first section:
        if device_category is not None:
            cell = ws.cell(row_number, 1, f'{device_category} Only')
            cell.font = bold_font
            for pattern_column_number in range(1, 9):
                ws.cell(row_number, pattern_column_number).fill = dark_grey_fill
            row_number += 1

        # Add the ECR to cat_df
        cat_df = cat_df.copy()
        cat_df['ecr'] = cat_df['transactions'] / cat_df['sessions']

        # Iterate through numerical data columns
        for column in ['transactions', 'sessions', 'ecr', 'quantity']:

            # Add subsection label
            section_label = column.capitalize() if column != 'ecr' else 'ECR'
            cell = ws.cell(row_number, 1, value=section_label)
            cell.font = bold_font
            for pattern_column_number in range(1, 9):
                ws.cell(row_number, pattern_column_number).fill = light_grey_fill
            row_number += 1

            # Iterate through metrics
            ws.cell(row_number, 1, value='Mean')
            for column_number, day_of_week in enumerate(days_of_week, start=2):
                ws.cell(
                    row_number, column_number,
                    value=cat_df[cat_df['day_of_week'] == day_of_week][column].mean()
                ).number_format = '0.##' if column != 'ecr' else '0.0%'
            row_number += 1
            ws.cell(row_number, 1, value='Std Dev')
            for column_number, day_of_week in enumerate(days_of_week, start=2):
                ws.cell(
                    row_number, column_number,
                    value=cat_df[cat_df['day_of_week'] == day_of_week][column].std()
                ).number_format = '0.##' if column != 'ecr' else '0.0%'
            row_number += 1
            ws.cell(row_number, 1, value='Min')
            for column_number, day_of_week in enumerate(days_of_week, start=2):
                ws.cell(
                    row_number, column_number,
                    value=cat_df[cat_df['day_of_week'] == day_of_week][column].min()
                ).number_format = '###,###,###' if column != 'ecr' else '0.0%'
            row_number += 1
            ws.cell(row_number, 1, value='1st Quartile')
            for column_number, day_of_week in enumerate(days_of_week, start=2):
                ws.cell(
                    row_number, column_number,
                    value=cat_df[cat_df['day_of_week'] == day_of_week][column].quantile(.25)
                ).number_format = '###,###,###' if column != 'ecr' else '0.0%'
            row_number += 1
            ws.cell(row_number, 1, value='Median')
            for column_number, day_of_week in enumerate(days_of_week, start=2):
                ws.cell(
                    row_number, column_number,
                    value=cat_df[cat_df['day_of_week'] == day_of_week][column].median()
                ).number_format = '###,###,###' if column != 'ecr' else '0.0%'
            row_number += 1
            ws.cell(row_number, 1, value='3rd Quartile')
            for column_number, day_of_week in enumerate(days_of_week, start=2):
                ws.cell(
                    row_number, column_number,
                    value=cat_df[cat_df['day_of_week'] == day_of_week][column].quantile(.75)
                ).number_format = '###,###,###' if column != 'ecr' else '0.0%'
            row_number += 1
            ws.cell(row_number, 1, value='Max')
            for column_number, day_of_week in enumerate(days_of_week, start=2):
                ws.cell(
                    row_number, column_number,
                    value=cat_df[cat_df['day_of_week'] == day_of_week][column].max()
                ).number_format = '###,###,###' if column != 'ecr' else '0.0%'
            row_number += 1
            ws.cell(row_number, 1, value='IQR')
            for column_number, day_of_week in enumerate(days_of_week, start=2):
                ws.cell(
                    row_number, column_number,
                    value=(
                            cat_df[cat_df['day_of_week'] == day_of_week][column].quantile(.75) -
                            cat_df[cat_df['day_of_week'] == day_of_week][column].quantile(.25)
                    )
                ).number_format = '###,###,###' if column != 'ecr' else '0.0%'
            row_number += 1

        row_number += 2

    # Set first column width and freeze 1st row and 1st column
    ws.column_dimensions['A'].width = 20
    ws.freeze_panes = 'B2'

    # I'm going to throw some pertinent figures in as well. Just as images.
    df = sessions_df.copy()
    df = sessions_df.groupby(['date', 'device_category'], as_index=False).sum()
    df['ecr'] = df['transactions'] / df['sessions']
    day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    df['day_of_week'] = pd.Categorical(df['date'].dt.day_name(), categories=day_order, ordered=True)

    # Transactions:
    fig, ax = plt.subplots()  # type: plt.Figure, plt.Axes
    sns.lineplot(
        x='day_of_week', y='transactions', hue='device_category', data=df,
        estimator='mean', errorbar=('ci', 95), ax=ax,
        style='device_category', markers={'Tablet': 'o', 'Desktop': 's', 'Mobile': 'D'}
    )
    fig.set_size_inches(11, 6.8)
    fig.set_dpi(300)
    title = (f"Mean Transactions by Day of Week and Device Category\n"
             f"With 95% Bootstrapped Confidence Intervals")
    ax.set_title(title, fontsize=20)
    ax.set_xlabel('Day of Week', fontsize=16)
    ax.tick_params(axis='x', labelsize=12)
    ax.set_ylabel(f"Mean Transactions", fontsize=16)
    ax.tick_params(axis='y', labelsize=12)
    ax.yaxis.set_major_formatter(FuncFormatter(lambda x, p: format(int(x), ',')))
    ax.legend(title='Device Category', loc='upper left', title_fontsize=12, fontsize=12)
    img_filepath = os.path.join(xlsx_figs_dirpath, 'day_of_week_transactions_img.png')
    fig.savefig(img_filepath)
    plt.close()
    ws.add_image(Image(img_filepath), 'J3')

    # Sessions
    fig, ax = plt.subplots()  # type: plt.Figure, plt.Axes
    sns.lineplot(
        x='day_of_week', y='sessions', hue='device_category', data=df,
        estimator='mean', errorbar=('ci', 95), ax=ax,
        style='device_category', markers={'Tablet': 'o', 'Desktop': 's', 'Mobile': 'D'}
    )
    fig.set_size_inches(11, 6.8)
    fig.set_dpi(300)
    title = (f"Mean Sessions by Day of Week and Device Category\n"
             f"With 95% Bootstrapped Confidence Intervals")
    ax.set_title(title, fontsize=20)
    ax.set_xlabel('Day of Week', fontsize=16)
    ax.tick_params(axis='x', labelsize=12)
    ax.set_ylabel(f"Mean Sessions", fontsize=16)
    ax.tick_params(axis='y', labelsize=12)
    ax.yaxis.set_major_formatter(FuncFormatter(lambda x, p: format(int(x), ',')))
    ax.legend(title='Device Category', loc='upper left', title_fontsize=12, fontsize=12)
    img_filepath = os.path.join(xlsx_figs_dirpath, 'day_of_week_sessions_img.png')
    fig.savefig(img_filepath)
    plt.close()
    ws.add_image(Image(img_filepath), 'J42')

    # ECR
    fig, ax = plt.subplots()  # type: plt.Figure, plt.Axes
    sns.lineplot(
        x='day_of_week', y='ecr', hue='device_category', data=df,
        estimator='mean', errorbar=('ci', 95), ax=ax,
        style='device_category', markers={'Tablet': 'o', 'Desktop': 's', 'Mobile': 'D'}
    )
    fig.set_size_inches(11, 6.8)
    fig.set_dpi(300)
    title = (f"Mean ECR by Day of Week and Device Category\n"
             f"With 95% Bootstrapped Confidence Intervals")
    ax.set_title(title, fontsize=20)
    ax.set_xlabel('Day of Week', fontsize=16)
    ax.tick_params(axis='x', labelsize=12)
    ax.set_ylim(bottom=0.05, top=0.045)
    ax.set_ylabel(f"Mean ECR (%)", fontsize=16)
    ax.tick_params(axis='y', labelsize=12)
    ax.yaxis.set_major_formatter(PercentFormatter(xmax=1, symbol=''))
    ax.legend(title='Device Category', loc='upper left', title_fontsize=12, fontsize=12)
    img_filepath = os.path.join(xlsx_figs_dirpath, 'day_of_week_ecr_img.png')
    fig.savefig(img_filepath)
    plt.close()
    ws.add_image(Image(img_filepath), 'J81')

    # Quantity
    fig, ax = plt.subplots()  # type: plt.Figure, plt.Axes
    sns.lineplot(
        x='day_of_week', y='transactions', hue='device_category', data=df,
        estimator='mean', errorbar=('ci', 95), ax=ax,
        style='device_category', markers={'Tablet': 'o', 'Desktop': 's', 'Mobile': 'D'}
    )
    fig.set_size_inches(11, 6.8)
    fig.set_dpi(300)
    title = (f"Mean Quantity by Day of Week and Device Category\n"
             f"With 95% Bootstrapped Confidence Intervals")
    ax.set_title(title, fontsize=20)
    ax.set_xlabel('Day of Week', fontsize=16)
    ax.tick_params(axis='x', labelsize=12)
    ax.set_ylabel(f"Mean Quantity", fontsize=16)
    ax.tick_params(axis='y', labelsize=12)
    ax.yaxis.set_major_formatter(FuncFormatter(lambda x, p: format(int(x), ',')))
    ax.legend(title='Device Category', loc='upper left', title_fontsize=12, fontsize=12)
    img_filepath = os.path.join(xlsx_figs_dirpath, 'day_of_week_quantity_img.png')
    fig.savefig(img_filepath)
    plt.close()
    ws.add_image(Image(img_filepath), 'J120')


if __name__ == '__main__':
    main()
